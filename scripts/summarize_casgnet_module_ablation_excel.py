#!/usr/bin/env python3
"""
汇总 CASGNet SA/GRN/SK 三模块 2^3 消融的 result_summary.json 到 Excel。

指标格式与 comparison_summary*.csv 一致：bootstrap n=1000 的 95%% CI，
形如 0.950(0.928-0.967)（优先用 bootstrap_auc / bootstrap_metrics 的 mean 与 ci95_low/high）。
"""
from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

METRIC_KEYS = ("auc", "sensitivity", "specificity", "npv", "ppv", "acc")

AB_DESC = {
    "000": "基线：无 SA、无 GRN、末 stage 无 SK（全 ASGBlock 且 SA/GRN 关闭）",
    "100": "仅 SA",
    "010": "仅 GRN",
    "001": "仅末 stage SK",
    "110": "SA + GRN",
    "101": "SA + SK",
    "011": "GRN + SK",
    "111": "完整 CASGNet-S1（SA+GRN+SK）",
}


def fmt_ci(mean_v, low_v, high_v) -> str:
    """与 refresh_supcon_checkpoint_metrics.fmt_ci / comparison_summary.csv 一致。"""
    if mean_v is None:
        return ""
    try:
        m = float(mean_v)
    except (TypeError, ValueError):
        return ""
    if low_v is None or high_v is None:
        return f"{m:.3f}"
    try:
        lo = float(low_v)
        hi = float(high_v)
    except (TypeError, ValueError):
        return f"{m:.3f}"
    return f"{m:.3f}({lo:.3f}-{hi:.3f})"


def metric_ci_strings(split: dict | None) -> dict[str, str]:
    """从 result_summary 或 test_eval 块提取带 bootstrap CI 的指标字符串。"""
    out: dict[str, str] = {k: "" for k in METRIC_KEYS}
    if not split:
        return out

    b_auc = split.get("bootstrap_auc") if isinstance(split.get("bootstrap_auc"), dict) else {}
    out["auc"] = fmt_ci(
        b_auc.get("mean", split.get("auc")),
        b_auc.get("ci95_low"),
        b_auc.get("ci95_high"),
    )

    bm = split.get("bootstrap_metrics") if isinstance(split.get("bootstrap_metrics"), dict) else {}
    for k in METRIC_KEYS:
        if k == "auc":
            continue
        block = bm.get(k) if isinstance(bm.get(k), dict) else {}
        out[k] = fmt_ci(
            block.get("mean", split.get(k)),
            block.get("ci95_low"),
            block.get("ci95_high"),
        )
    return out


def auc_sort_value(split: dict | None) -> float:
    if not split:
        return -1.0
    b = split.get("bootstrap_auc") if isinstance(split.get("bootstrap_auc"), dict) else {}
    v = b.get("mean", split.get("auc"))
    try:
        return float(v)
    except (TypeError, ValueError):
        return -1.0


def parse_ab_code(model: str) -> tuple[str, int, int, int] | None:
    m = re.search(r"_ab(\d{3})$", model)
    if not m:
        return None
    bits = m.group(1)
    return bits, int(bits[0]), int(bits[1]), int(bits[2])


def collect_rows(root: Path, run_leaf_suffix: str) -> list[dict]:
    rows: list[dict] = []
    for d in sorted(root.iterdir()):
        if not d.is_dir() or d.name in ("logs", "_status"):
            continue
        rs = d / "result_summary.json"
        if not rs.is_file():
            continue
        data = json.loads(rs.read_text(encoding="utf-8"))
        model = str(data.get("model", d.name.replace(run_leaf_suffix, "")))
        parsed = parse_ab_code(model)
        if parsed:
            bits, sa, grn, sk = parsed
            desc = AB_DESC.get(bits, "")
        else:
            bits, sa, grn, sk, desc = "", -1, -1, -1, ""

        val_m = metric_ci_strings(data)
        te_raw = data.get("test_eval") if isinstance(data.get("test_eval"), dict) else None
        te_m = metric_ci_strings(te_raw)

        row: dict = {
            "run_leaf": d.name,
            "model": model,
            "ab_bits": bits,
            "use_sa": sa,
            "use_grn": grn,
            "use_sk_last": sk,
            "description": desc,
            "run_dir": str(d.resolve()),
            "_val_auc_sort": auc_sort_value(data),
        }
        for k in METRIC_KEYS:
            row[f"val_{k}"] = val_m[k]
        for k in METRIC_KEYS:
            row[f"test_{k}"] = te_m[k]
        rows.append(row)

    rows.sort(key=lambda r: (-r["_val_auc_sort"], r["model"]))
    for r in rows:
        r.pop("_val_auc_sort", None)
    return rows


def write_comparison_csv(rows: list[dict], path: Path, prefix: str) -> None:
    """写出与 comparison_summary 列名一致的 CSV（仅指标列 + model）。"""
    fields = ["model"] + list(METRIC_KEYS)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(f"{prefix}{k}", "") if k != "model" else r["model"] for k in fields})


def write_excel(rows: list[dict], out: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "ablation"
    headers = [
        "model",
        "ab_bits",
        "use_sa",
        "use_grn",
        "use_sk_last",
        "description",
    ]
    for prefix in ("val_", "test_"):
        for k in METRIC_KEYS:
            headers.append(f"{prefix}{k}")
    headers.append("run_dir")
    ws.append(headers)
    for r in rows:
        ws.append(
            [
                r["model"],
                r["ab_bits"],
                r["use_sa"],
                r["use_grn"],
                r["use_sk_last"],
                r["description"],
            ]
            + [r[f"val_{k}"] for k in METRIC_KEYS]
            + [r[f"test_{k}"] for k in METRIC_KEYS]
            + [r["run_dir"]]
        )
    for col in range(1, len(headers) + 1):
        w = 36 if col == 6 else (48 if col == len(headers) else 18)
        ws.column_dimensions[get_column_letter(col)].width = w

    meta = wb.create_sheet("legend")
    meta.append(["说明", "指标格式 mean(ci95_low-ci95_high)，三位小数"])
    meta.append(["bootstrap", "来自 result_summary.json 中 n_bootstrap=1000 的 bootstrap_auc / bootstrap_metrics"])
    meta.append(["ab_bits", "位序(左→右): SA, GRN, 末stage_SK；1=开启"])
    for b, t in sorted(AB_DESC.items()):
        meta.append([b, t])

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def main() -> None:
    ap = argparse.ArgumentParser(
        description="汇总 CASGNet SA/GRN/SK 消融；指标含 bootstrap1000 的 95%% CI。"
    )
    ap.add_argument(
        "--experiment-root",
        type=Path,
        default=Path("checkpoints/casgnet_sa_grn_sk_ablation"),
        help="含各 casgnet_s1_ab*_ce_only/result_summary.json 的根目录",
    )
    ap.add_argument("--run-leaf-suffix", type=str, default="_ce_only")
    ap.add_argument("--output", type=Path, default=None)
    ap.add_argument(
        "--also-csv",
        action="store_true",
        help="同时写出 comparison_summary_val.csv / comparison_summary_test.csv（仅 CASGNet 消融行）",
    )
    args = ap.parse_args()
    root = args.experiment_root.resolve()
    if not root.is_dir():
        _repo = Path(__file__).resolve().parent.parent
        alt = (_repo / "checkpoints/starnetsk_sk_kernel_ablation").resolve()
        msg = f"目录不存在: {root}"
        if alt.is_dir():
            msg += (
                f"\n若当时 RUN_TAG/OUT_ROOT 与 StarNet-SK 消融相同，请改用:\n"
                f"  python3 scripts/summarize_casgnet_module_ablation_excel.py "
                f"--experiment-root {alt} --run-leaf-suffix {args.run_leaf_suffix}"
            )
        raise SystemExit(msg)

    out = args.output or (root / "summary_casgnet_sa_grn_sk_ablation.xlsx")
    out = out.resolve()
    rows = collect_rows(root, args.run_leaf_suffix)
    if not rows:
        raise SystemExit(f"{root} 下无 result_summary.json（需含 bootstrap_auc / bootstrap_metrics）")

    write_excel(rows, out)
    print(f"写入 {out} ，共 {len(rows)} 条（指标含 bootstrap 95% CI）。")

    if args.also_csv:
        val_csv = root / "comparison_summary_casgnet_ablation_val.csv"
        test_csv = root / "comparison_summary_casgnet_ablation_test.csv"
        write_comparison_csv(rows, val_csv, "val_")
        write_comparison_csv(rows, test_csv, "test_")
        print(f"  val  -> {val_csv}")
        print(f"  test -> {test_csv}")


if __name__ == "__main__":
    main()
