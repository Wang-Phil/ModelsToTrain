#!/usr/bin/env python3
"""
汇总 StarNet-SK（models/starnetsk.py）SK 核尺度消融实验到 Excel。

优先读取与 old_data_supcon_compare_v3 一致的 result_summary.json（train_casgnet_contrastive_newdata.py）；
若无则回退 train_multiclass.py 的 history.json。
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

MODEL_TO_SK_PAIR: dict[str, tuple[int, int]] = {
    "starnet_s1_sk13": (1, 3),
    "starnet_s1_sk15": (1, 5),
    "starnet_s1_sk17": (1, 7),
    "starnet_s1_sk19": (1, 9),
    "starnet_s1_sk35": (3, 5),
    "starnet_s1_sk37": (3, 7),
    "starnet_s1_sk39": (3, 9),
    "starnet_s1_sk57": (5, 7),
    "starnet_s1_sk59": (5, 9),
    "starnet_s1_sk79": (7, 9),
}


def _strip_suffix(name: str, suffix: str) -> str:
    if suffix and name.endswith(suffix):
        return name[: -len(suffix)]
    return name


def _best_from_history(history: dict) -> tuple[float, int, int]:
    accs = history.get("val_acc") or []
    if not accs:
        return float("nan"), -1, 0
    best = max(accs)
    ep = int(accs.index(best)) + 1
    return float(best), ep, len(accs)


def collect_rows(root: Path, run_leaf_suffix: str) -> list[dict]:
    rows: list[dict] = []
    for d in sorted(root.iterdir()):
        if not d.is_dir() or d.name in ("logs", "_status"):
            continue
        base_model = _strip_suffix(d.name, run_leaf_suffix)
        rs_path = d / "result_summary.json"
        hist_path = d / "history.json"

        k1, k2 = MODEL_TO_SK_PAIR.get(base_model, (-1, -1))
        row: dict = {
            "run_dir": str(d.resolve()),
            "run_leaf": d.name,
            "model": base_model,
            "sk_kernel_1": k1,
            "sk_kernel_2": k2,
            "sk_pair_str": f"[{k1}, {k2}]" if k1 > 0 else "",
            "val_auc": None,
            "val_acc": None,
            "test_auc": None,
            "test_acc": None,
            "best_epoch": None,
            "epochs_ran": None,
            "source": "",
        }

        if rs_path.is_file():
            with rs_path.open(encoding="utf-8") as f:
                data = json.load(f)
            row["model"] = str(data.get("model", row["model"]))
            row["val_auc"] = data.get("auc")
            row["val_acc"] = data.get("acc")
            te = data.get("test_eval") if isinstance(data.get("test_eval"), dict) else None
            if te:
                row["test_auc"] = te.get("auc")
                row["test_acc"] = te.get("acc")
            row["source"] = "result_summary.json"
            rows.append(row)
            continue

        if hist_path.is_file():
            with hist_path.open(encoding="utf-8") as f:
                history = json.load(f)
            best_acc, best_ep, n_ep = _best_from_history(history)
            row["val_acc"] = best_acc / 100.0 if best_acc > 1.0 else best_acc
            row["best_epoch"] = best_ep
            row["epochs_ran"] = n_ep
            row["source"] = "history.json"
            rows.append(row)
            continue

    def sort_key(r: dict) -> tuple:
        va = r.get("val_auc")
        if va is None:
            va = -1.0
        try:
            va = float(va)
        except (TypeError, ValueError):
            va = -1.0
        acc = r.get("val_acc")
        try:
            acc = float(acc) if acc is not None else -1.0
        except (TypeError, ValueError):
            acc = -1.0
        return (-va, -acc, r["model"])

    rows.sort(key=sort_key)
    return rows


def write_excel(rows: list[dict], out_xlsx: Path, run_leaf_suffix: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "sk_kernel_ablation"
    headers = [
        "model",
        "sk_kernel_1",
        "sk_kernel_2",
        "sk_pair_str",
        "val_auc",
        "val_acc",
        "test_auc",
        "test_acc",
        "best_epoch",
        "epochs_ran",
        "source",
        "run_dir",
    ]
    ws.append(headers)
    for r in rows:
        ws.append(
            [
                r["model"],
                r["sk_kernel_1"],
                r["sk_kernel_2"],
                r["sk_pair_str"],
                r["val_auc"],
                r["val_acc"],
                r["test_auc"],
                r["test_acc"],
                r["best_epoch"],
                r["epochs_ran"],
                r["source"],
                r["run_dir"],
            ]
        )
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14 if col < 12 else 48
    meta = wb.create_sheet("meta")
    meta.append(["run_leaf_suffix", run_leaf_suffix])
    meta.append(["note", "val_* / test_* 来自 result_summary.json（与 old_data_supcon_compare_v3 一致）"])
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)


def main() -> None:
    ap = argparse.ArgumentParser(description="汇总 starnetsk SK 核消融到 Excel（优先 result_summary.json）")
    ap.add_argument(
        "--experiment-root",
        type=Path,
        default=Path("checkpoints/starnetsk_sk_kernel_supcon"),
        help="对比根目录（与 RUN_TAG 输出一致）",
    )
    ap.add_argument(
        "--output",
        type=Path,
        default=None,
        help="输出 .xlsx；默认 <experiment-root>/summary_sk_kernel_ablation.xlsx",
    )
    ap.add_argument(
        "--run-leaf-suffix",
        type=str,
        default="_ce_only",
        help="子目录名后缀，用于从目录名还原 --model（默认与 run_compare 一致 _ce_only）",
    )
    args = ap.parse_args()
    root = args.experiment_root.resolve()
    if not root.is_dir():
        raise SystemExit(f"目录不存在: {root}")

    out = args.output
    if out is None:
        out = root / "summary_sk_kernel_ablation.xlsx"
    else:
        out = out.resolve()

    rows = collect_rows(root, args.run_leaf_suffix)
    if not rows:
        raise SystemExit(f"在 {root} 下未找到含 result_summary.json 或 history.json 的子目录。")

    write_excel(rows, out, args.run_leaf_suffix)
    print(f"写入 {out} ，共 {len(rows)} 条。")


if __name__ == "__main__":
    main()
