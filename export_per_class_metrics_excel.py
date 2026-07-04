#!/usr/bin/env python3
"""
将 analyze_supcon_per_class_metrics.py 生成的 JSON 导出为 Excel（.xlsx）。

依赖: openpyxl
  pip install openpyxl

用法:
  python export_per_class_metrics_excel.py \\
    --input-json checkpoints/casgnet_supcon_newdata/per_class_metrics_val.json \\
    --output-xlsx checkpoints/casgnet_supcon_newdata/per_class_metrics_val.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any


def write_per_class_metrics_excel(payload: dict[str, Any], xlsx_path: str | Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as e:
        raise SystemExit("请先安装: pip install openpyxl") from e

    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # --- Sheet: 各类别指标 ---
    ws1 = wb.active
    ws1.title = "各类别指标"
    per = payload.get("per_class") or []
    if not per:
        ws1.append(["(无 per_class 数据)"])
    else:
        keys = [
            "class",
            "class_idx",
            "support_true_k",
            "support_pred_k",
            "TP",
            "FP",
            "FN",
            "TN",
            "AUC",
            "Sensitivity",
            "Specificity",
            "PPV",
            "NPV",
            "ACC_ovr",
        ]
        headers_cn = [
            "类别",
            "类别索引",
            "真实该类样本数",
            "预测为该类次数",
            "TP",
            "FP",
            "FN",
            "TN",
            "AUC",
            "灵敏度 Sensitivity",
            "特异度 Specificity",
            "阳性预测值 PPV",
            "阴性预测值 NPV",
            "OvR二分类准确率 ACC",
        ]
        ws1.append(headers_cn)
        for row in per:
            ws1.append([row.get(k) for k in keys])
        for cell in ws1[1]:
            cell.font = Font(bold=True)
        for col in ws1.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws1.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 45)
        for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, min_col=9, max_col=14):
            for c in row:
                c.number_format = "0.0000"

    # --- Sheet: 汇总 ---
    ws2 = wb.create_sheet("汇总与说明", 0)
    meta_keys = ["checkpoint", "data_dir", "split", "model"]
    ws2.append(["字段", "值"])
    for k in meta_keys:
        ws2.append([k, str(payload.get(k, ""))])
    ws2.append([])
    summ = payload.get("summary") or {}
    ws2.append(["汇总指标", "值"])
    for k, v in summ.items():
        ws2.append([k, v if isinstance(v, (int, str)) else float(v)])
    ws2.append([])
    notes = payload.get("notes") or {}
    ws2.append(["说明", ""])
    for k, v in notes.items():
        ws2.append([k, str(v)])
    ws2["A1"].font = Font(bold=True)
    ws2["A1"].alignment = Alignment(wrap_text=True)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 88

    wb.save(str(xlsx_path))


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input-json", type=str, required=True)
    ap.add_argument("--output-xlsx", type=str, required=True)
    args = ap.parse_args()
    jp = Path(args.input_json)
    if not jp.is_file():
        print(f"找不到 JSON: {jp}", file=sys.stderr)
        sys.exit(1)
    with open(jp, encoding="utf-8") as f:
        payload = json.load(f)
    write_per_class_metrics_excel(payload, args.output_xlsx)
    print(f"已写入: {Path(args.output_xlsx).resolve()}")


if __name__ == "__main__":
    main()
